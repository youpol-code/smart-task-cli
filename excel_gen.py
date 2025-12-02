from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font,PatternFill
from typing import cast,List 

def create_report() -> None:
    wb: Workbook = Workbook()

    ws: Worksheet = cast(Worksheet,wb.active)
    ws.title ="Sales Report"

    headers: List[str] =["รายการสินค้า","ราคา (บาท)","จำนวน (ชิ้่น)","ราคารวม (บาท)"]
    ws.append(headers)

    data = [
        ["คีย์บอร์ดไร้สาย",1500,2],
        ["เมาส์เพื่อสุขภาพ",950,5],
        ["แผ่นรองเมาร์",120,10]
    ]

    for item in data:
        total_price = item[1] * item[2]
        full_row =[item[0],item[1],item[2],total_price]
        ws.append(full_row)

    header_font = Font(bold=True,color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD",end_color="4F81BD",fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['D'].width = 15

    filename="report_output.xlsx"
    wb.save(filename)
    print(f"สร้างไฟล์ {filename} เสร็จสมบูรณ์ (ลองเปิดดูสิ สวยไหม?)!")

if __name__ =="__main__":
    create_report()